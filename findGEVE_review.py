#!/usr/bin/env python3
"""
findGEVE_review - Review findGEVE calls and annotate manually specified candidate regions.
Author: Dede Kurniawan (dedekurniawan@genomics.cn)
"""

from __future__ import annotations

import argparse
import gzip
import logging
import math
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    raise SystemExit(
        "Error: openpyxl is required for review.xlsx support. "
        "Install with: pip install openpyxl"
    ) from exc

HELP_TEXT = """\
{prog} - Review findGEVE results or annotate manually selected regions.

Usage: {prog} <command> [OPTIONS]

Commands:
  make-template   Create a review.xlsx workbook from findGEVE summary output
  apply           Apply review decisions and reannotate changed GEVE regions
  region          Annotate one or more exact manually selected candidate regions

Run {prog} <command> --help for command-specific options.
"""

MAKE_TEMPLATE_HELP_TEXT = """\
{prog} make-template - Create a review workbook.

Usage: {prog} make-template <prefix.summary.tsv> [OPTIONS]

Mandatory:
  summary              Input findGEVE <prefix>.summary.tsv

Optionals:
  --overwrite          Replace an existing <prefix>.review.xlsx
  -h, --help           Show this help and exit
"""

APPLY_HELP_TEXT = """\
{prog} apply - Apply manual review decisions to findGEVE results.

Usage:
  {prog} apply \\
    -db <directory> \\
    --genome genome.fa \\
    --review <prefix.review.xlsx> \\
    --summary <prefix.summary.tsv> \\
    --markerout <prefix.markerout> [OPTIONS]

Mandatory:
  -db, --db            HMM database directory containing NCLDV_markers.hmm
                       and gvog.complete.hmm; Pfam-A.hmm is optional
  --genome             Input genome assembly FASTA; gzip is acceptable
  --review             Reviewed <prefix>.review.xlsx workbook
  --summary            Original findGEVE <prefix>.summary.tsv
  --markerout          Original findGEVE <prefix>.markerout

Optionals:
  -g, --gff            Host eukaryotic annotation in GFF/GFF3 format
  --original-pep       Original findGEVE <prefix>.geve.pep; auto-detected
                       beside --summary when omitted
  --prefix             Output prefix inferred from --summary when omitted
  -o, --outdir         Output directory
  --outbase            Base directory for an automatic Review_<YYYYMMDD> folder
  -t, --threads        CPU threads for ORF prediction and HMM search [default: 4]
  -e, --evalue         E-value cutoff for HMM searches              [default: 1e-5]
  --overwrite          Allow writing into a non-empty output directory
  --no-plot            Skip automatic plotting
  -h, --help           Show this help and exit

Review actions:
  unchanged            Keep the original GEVE call and annotation
  remove               Remove the GEVE from reviewed outputs
  change               Use exact reviewed coordinates and reannotate the full region
"""

REGION_HELP_TEXT = """\
{prog} region - Annotate exact manually selected candidate regions.

Usage:
  {prog} region \\
    -db <directory> --genome genome.fa --prefix <prefix> \\
    --ctg <contig[,contig...]> \\
    --start <start[,start...]> --end <end[,end...]> [OPTIONS]

Mandatory:
  -db, --db            HMM database directory containing NCLDV_markers.hmm
                       and gvog.complete.hmm; Pfam-A.hmm is optional
  --genome             Input genome assembly FASTA; gzip is acceptable
  --prefix             Output prefix and GEVE ID prefix
  --ctg                One or more contigs separated by commas
  --start              Matching 1-based inclusive starts separated by commas
  --end                Matching 1-based inclusive ends separated by commas

Optionals:
  -g, --gff            Host eukaryotic annotation in GFF/GFF3 format
  -o, --outdir         Output directory
  --outbase            Base directory for an automatic Result_<YYYYMMDD> folder
  -t, --threads        CPU threads for ORF prediction and HMM search [default: 4]
  -e, --evalue         E-value cutoff for HMM searches              [default: 1e-5]
  --overwrite          Allow writing into a non-empty output directory
  --no-plot            Skip automatic plotting
  -h, --help           Show this help and exit

Examples:
  {prog} region -db DB --genome genome.fa --prefix sample \\
    --ctg ctg1 --start 1 --end 10000

  {prog} region -db DB --genome genome.fa --prefix sample \\
    --ctg ctg1,ctg2 --start 1,500 --end 10000,20000

The three lists must contain the same number of values. Candidate IDs follow
input order: <prefix>_GEVE_001, <prefix>_GEVE_002, and so on. Every exact
interval is retained even when no hallmark gene is detected.
"""

_LOG = logging.getLogger("findGEVE_review")
OUTPUT = 25
logging.addLevelName(OUTPUT, "OUTPUT")

def _output(self, message, *args, **kwargs):
    if self.isEnabledFor(OUTPUT):
        self._log(OUTPUT, message, args, **kwargs)

logging.Logger.output = _output

ACTIONS = ("unchanged", "remove", "change")
REVIEW_COLUMNS = [
    "geve_name",
    "action",
    "contig",
    "original_start",
    "original_end",
    "review_start",
    "review_end",
]

FEATURE_IGNORE = {
    "GEVE",
    "flank_left",
    "flank_right",
    "TIR_left",
    "TIR_right",
    "TSD_5p",
    "TSD_3p",
}

BLASTN_COLUMNS = [
    "qstart", "qend", "sstart", "send", "length", "nident",
    "pident", "gaps", "evalue", "bitscore",
]
BLASTN_OUTFMT = "6 " + " ".join(BLASTN_COLUMNS)
NATKEY_SPLIT = re.compile(r"(\d+)")
REVCOMP_TABLE = str.maketrans("ACGTacgtNn", "TGCAtgcaNn")
CODON_TABLE = {
    "TTT":"F","TTC":"F","TTA":"L","TTG":"L","TCT":"S","TCC":"S","TCA":"S","TCG":"S",
    "TAT":"Y","TAC":"Y","TAA":"*","TAG":"*","TGT":"C","TGC":"C","TGA":"*","TGG":"W",
    "CTT":"L","CTC":"L","CTA":"L","CTG":"L","CCT":"P","CCC":"P","CCA":"P","CCG":"P",
    "CAT":"H","CAC":"H","CAA":"Q","CAG":"Q","CGT":"R","CGC":"R","CGA":"R","CGG":"R",
    "ATT":"I","ATC":"I","ATA":"I","ATG":"M","ACT":"T","ACC":"T","ACA":"T","ACG":"T",
    "AAT":"N","AAC":"N","AAA":"K","AAG":"K","AGT":"S","AGC":"S","AGA":"R","AGG":"R",
    "GTT":"V","GTC":"V","GTA":"V","GTG":"V","GCT":"A","GCC":"A","GCA":"A","GCG":"A",
    "GAT":"D","GAC":"D","GAA":"E","GAG":"E","GGT":"G","GGC":"G","GGA":"G","GGG":"G",
}

@dataclass
class TirPair:
    left_start: int
    left_end: int
    right_start: int
    right_end: int
    tir_length: int
    insert_size: int
    tir_identity: float
    score: int
    matches: int
    total: int
    gaps: int
    tir_evalue: float = float("nan")

@dataclass
class Tsd:
    sequence_left: str
    sequence_right: str
    length: int
    mismatches: int
    identity: float
    left_shift: int
    right_shift: int

def setup_logging(log_path: Optional[Path] = None) -> None:
    _LOG.setLevel(logging.DEBUG)
    _LOG.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    sh = logging.StreamHandler(sys.stderr)
    sh.setFormatter(fmt)
    sh.setLevel(logging.INFO)
    _LOG.addHandler(sh)
    if log_path is not None:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(log_path, mode="w")
        fh.setFormatter(fmt)
        fh.setLevel(logging.DEBUG)
        _LOG.addHandler(fh)

def _natural_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in NATKEY_SPLIT.split(str(s))]

def _require_columns(df: pd.DataFrame, cols: Iterable[str], label: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise SystemExit(f"Error: {label} is missing required column(s): {', '.join(missing)}")

def _read_table(path: Path, label: str) -> pd.DataFrame:
    if path is None or not path.is_file():
        raise SystemExit(f"Error: {label} file not found: {path}")
    try:
        return pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False)
    except Exception as exc:
        raise SystemExit(f"Error: failed to read {label}: {path}: {exc}") from exc

def _safe_int(value, default: Optional[int] = None) -> Optional[int]:
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    s = str(value).strip()
    if not s or s.upper() in {"NA", "NAN", "NONE"}:
        return default
    try:
        return int(float(s.replace(",", "")))
    except ValueError:
        return default

def _safe_float(value, default: float = float("nan")) -> float:
    if value is None:
        return default
    s = str(value).strip()
    if not s or s.upper() in {"NA", "NAN", "NONE"}:
        return default
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return default

def _normalize_action(value) -> str:
    return str(value or "").strip().lower()

def infer_prefix(summary_path: Path, summary: Optional[pd.DataFrame] = None) -> str:
    name = summary_path.name
    for suffix in (".summary.tsv", ".tsv", ".txt"):
        if name.endswith(suffix):
            name = name[:-len(suffix)]
            break
    if name:
        return name
    if summary is not None and "geve_name" in summary.columns and not summary.empty:
        first = str(summary["geve_name"].iloc[0])
        return first.split("_GEVE_", 1)[0] if "_GEVE_" in first else first.split("_", 1)[0]
    return "findGEVE"

def default_outdir(base: Optional[Path]) -> Path:
    root = base if base is not None else Path.cwd()
    date_tag = datetime.now().strftime("%Y%m%d")
    out = root / f"Review_{date_tag}"
    if not out.exists():
        return out
    idx = 1
    while True:
        cand = root / f"Review_{date_tag}_{idx:02d}"
        if not cand.exists():
            return cand
        idx += 1

def read_fasta(path: Optional[Path]) -> Dict[str, str]:
    if path is None:
        return {}
    if not path.is_file():
        raise SystemExit(f"Error: genome FASTA not found: {path}")
    opener = gzip.open if str(path).endswith(".gz") else open
    seqs: Dict[str, List[str]] = {}
    current: Optional[str] = None
    try:
        with opener(path, "rt") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                if line.startswith(">"):
                    current = line[1:].split()[0]
                    seqs.setdefault(current, [])
                elif current is not None:
                    seqs[current].append(line.upper())
    except Exception as exc:
        raise SystemExit(f"Error: failed to read genome FASTA {path}: {exc}") from exc
    return {k: "".join(v) for k, v in seqs.items()}

def fetch_seq(seqs: Dict[str, str], contig: str, start: int, end: int) -> str:
    seq = seqs.get(str(contig))
    if seq is None:
        return ""
    start = max(1, int(start))
    end = min(len(seq), int(end))
    if end < start:
        return ""
    return seq[start - 1:end]

def gc_of_seq(seq: str) -> float:
    s = seq.upper()
    gc = s.count("G") + s.count("C")
    at = s.count("A") + s.count("T")
    return float(100.0 * gc / (gc + at)) if (gc + at) else float("nan")

def wrap_fasta(seq: str, width: int = 80) -> str:
    return "\n".join(seq[i:i + width] for i in range(0, len(seq), width))

def revcomp(seq: str) -> str:
    return seq.translate(REVCOMP_TABLE)[::-1]

def translate_cds(seq: str) -> str:
    seq = seq.upper().replace("U", "T")
    aa = []
    for i in range(0, len(seq) - 2, 3):
        codon = seq[i:i + 3]
        if any(b not in "ACGT" for b in codon):
            aa.append("X")
        else:
            aa.append(CODON_TABLE.get(codon, "X"))
    return "".join(aa).rstrip("*")

def viz_flank_size(geve_length: int) -> int:
    geve_length = max(1, int(geve_length))
    return int(min(200_000, max(10_000, round(geve_length * 0.10))))

def make_template(summary_path: Path, overwrite: bool = False) -> Path:
    summary = _read_table(summary_path, "summary")
    _require_columns(summary, ["geve_name", "start", "end"], "summary")
    contig_col = "contig_id" if "contig_id" in summary.columns else "contig"
    if contig_col not in summary.columns:
        raise SystemExit("Error: summary must contain contig_id or contig column")
    prefix = infer_prefix(summary_path, summary)
    out_path = summary_path.with_name(f"{prefix}.review.xlsx")
    if out_path.exists() and not overwrite:
        raise SystemExit(f"Error: output exists, use --overwrite to replace: {out_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    locked_fill = PatternFill("solid", fgColor="D9EAF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(REVIEW_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for _, row in summary.sort_values("geve_name", key=lambda c: c.map(_natural_key)).iterrows():
        ws.append([
            row.get("geve_name", ""),
            "unchanged",
            row.get(contig_col, ""),
            _safe_int(row.get("start", "")),
            _safe_int(row.get("end", "")),
            "",
            "",
        ])

    dv = DataValidation(type="list", formula1='"unchanged,remove,change"', allow_blank=False)
    dv.error = "Choose one of: unchanged, remove, change"
    dv.errorTitle = "Invalid action"
    dv.prompt = "Choose unchanged, remove, or change"
    dv.promptTitle = "GEVE review action"
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"B2:B{ws.max_row}")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            if idx in (2, 6, 7):
                cell.fill = editable_fill
                cell.protection = Protection(locked=False)
            else:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)

    widths = {"A": 28, "B": 14, "C": 12, "D": 16, "E": 16, "F": 16, "G": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.protection.sheet = True
    ws.protection.enable()

    note = wb.create_sheet("README")
    lines = [
        "findGEVE review.xlsx instructions",
        "Allowed action values: unchanged, remove, change",
        "unchanged: keep the original GEVE call. Leave review_start/review_end empty.",
        "remove: drop the GEVE from reviewed outputs. Leave review_start/review_end empty.",
        "change: fill review_start and/or review_end. Blank side uses the original coordinate.",
        "Use the Plotly HTML coordinate-review plot to click and copy boundary coordinates.",
    ]
    for line in lines:
        note.append([line])
    note["A1"].font = Font(bold=True, size=14)
    note.column_dimensions["A"].width = 120

    wb.save(out_path)
    _LOG.info(f"Wrote review template: {out_path}")
    return out_path

def read_review_xlsx(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise SystemExit(f"Error: review file not found: {path}")
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        raise SystemExit(f"Error: failed to open review workbook {path}: {exc}") from exc
    if "review" not in wb.sheetnames:
        raise SystemExit("Error: review workbook must contain a sheet named 'review'")
    ws = wb["review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Error: review sheet is empty")
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header).dropna(how="all")
    _require_columns(df, REVIEW_COLUMNS, "review.xlsx")
    return df[REVIEW_COLUMNS].copy()

def validate_review(review: pd.DataFrame, summary: pd.DataFrame) -> Tuple[pd.DataFrame, List[str], List[str]]:
    _require_columns(summary, ["geve_name", "start", "end"], "summary")
    contig_col = "contig_id" if "contig_id" in summary.columns else "contig"
    if contig_col not in summary.columns:
        raise SystemExit("Error: summary must contain contig_id or contig column")
    known = {str(r["geve_name"]): r for _, r in summary.iterrows()}

    errors: List[str] = []
    warnings: List[str] = []
    clean_rows = []
    seen = set()

    for i, row in review.iterrows():
        excel_row = i + 2
        geve_name = str(row.get("geve_name", "")).strip()
        action = _normalize_action(row.get("action", ""))
        if not geve_name:
            errors.append(f"row {excel_row}: geve_name is empty")
            continue
        if geve_name in seen:
            errors.append(f"row {excel_row}: duplicate geve_name: {geve_name}")
            continue
        seen.add(geve_name)
        if geve_name not in known:
            errors.append(f"row {excel_row}: geve_name not present in summary: {geve_name}")
            continue
        if action not in ACTIONS:
            errors.append(
                f"row {excel_row} ({geve_name}): invalid action {row.get('action')!r}; "
                "must be unchanged, remove, or change"
            )
            continue

        expected_contig = str(known[geve_name].get(contig_col, "")).strip()
        contig = str(row.get("contig", "")).strip()
        if not contig:
            errors.append(f"row {excel_row} ({geve_name}): contig is empty in review.xlsx")
            continue
        if contig != expected_contig:
            errors.append(
                f"row {excel_row} ({geve_name}): contig {contig!r} does not match summary {expected_contig!r}"
            )

        orig_start = _safe_int(row.get("original_start"))
        orig_end = _safe_int(row.get("original_end"))
        if orig_start is None or orig_end is None:
            errors.append(f"row {excel_row} ({geve_name}): original_start/original_end must be numeric in review.xlsx")
            continue
        if orig_start >= orig_end:
            errors.append(f"row {excel_row} ({geve_name}): original_start must be smaller than original_end")
            continue

        review_start = _safe_int(row.get("review_start"))
        review_end = _safe_int(row.get("review_end"))
        has_review_bounds = review_start is not None or review_end is not None
        summary_start = _safe_int(known[geve_name].get("start"))
        summary_end = _safe_int(known[geve_name].get("end"))
        original_bounds_changed = (
            summary_start is not None and summary_end is not None
            and (orig_start != summary_start or orig_end != summary_end)
        )

        if action == "unchanged" and (has_review_bounds or original_bounds_changed):
            action = "change"
            warnings.append(
                f"row {excel_row} ({geve_name}): reviewed coordinates differ from the original call; "
                "treated as change"
            )

        if action == "change":
            if review_start is None:
                review_start = orig_start
            if review_end is None:
                review_end = orig_end
            if (
                not original_bounds_changed
                and not has_review_bounds
                and review_start == orig_start
                and review_end == orig_end
            ):
                errors.append(f"row {excel_row} ({geve_name}): change requires a modified coordinate in review.xlsx")
            elif review_start >= review_end:
                errors.append(f"row {excel_row} ({geve_name}): reviewed start must be smaller than reviewed end")
        elif has_review_bounds:
            warnings.append(f"row {excel_row} ({geve_name}): review_start/review_end ignored for action={action}")

        clean_rows.append(dict(
            geve_name=geve_name,
            action=action,
            contig=contig,
            original_start=orig_start,
            original_end=orig_end,
            review_start=review_start,
            review_end=review_end,
        ))

    missing = sorted(set(known) - seen, key=_natural_key)
    for geve_name in missing:
        errors.append(f"{geve_name}: missing from review.xlsx; keep the row and use action=remove to exclude it")

    clean = pd.DataFrame(clean_rows)
    if not clean.empty:
        clean = clean.reset_index(drop=True)
    return clean, errors, warnings

def parse_blastn_tabular(tab_path: Path) -> List[TirPair]:
    if not tab_path.exists() or not tab_path.read_text().strip():
        return []
    pairs: List[TirPair] = []
    for line in tab_path.read_text().splitlines():
        fields = line.rstrip("\n").split("\t")
        if len(fields) < 10:
            continue
        try:
            qstart, qend = int(fields[0]), int(fields[1])
            sstart, send = int(fields[2]), int(fields[3])
            aln_len, nident = int(fields[4]), int(fields[5])
            pident, gaps = float(fields[6]), int(fields[7])
            evalue, bitscore = float(fields[8]), float(fields[9])
        except ValueError:
            continue
        left_start, left_end = min(qstart, qend), max(qstart, qend)
        right_start, right_end = min(sstart, send), max(sstart, send)
        if left_end >= right_start:
            continue
        tir_length = left_end - left_start + 1
        insert_size = right_start - left_end - 1
        pairs.append(TirPair(left_start, left_end, right_start, right_end,
                             tir_length, insert_size, pident, int(round(bitscore)),
                             nident, aln_len, gaps, evalue))
    return pairs

def run_blastn_self(region_seq: str, cfg: dict, threads: int = 1) -> List[TirPair]:
    blastn = shutil.which("blastn")
    if blastn is None:
        raise RuntimeError("blastn executable was not found in PATH")
    with tempfile.TemporaryDirectory(prefix="findGEVE_review_tir_") as tmp:
        tmpdir = Path(tmp)
        fa_path = tmpdir / "candidate.fa"
        tab_path = tmpdir / "candidate.blastn.tsv"
        fa_path.write_text(">candidate\n" + wrap_fasta(region_seq) + "\n")
        cmd = [
            blastn, "-query", str(fa_path), "-subject", str(fa_path),
            "-strand", "minus", "-task", "blastn",
            "-word_size", str(cfg["blastn_word_size"]),
            "-reward", str(cfg["blastn_reward"]),
            "-penalty", str(cfg["blastn_penalty"]),
            "-gapopen", str(cfg["blastn_gapopen"]),
            "-gapextend", str(cfg["blastn_gapextend"]),
            "-evalue", str(cfg["blastn_evalue"]),
            "-dust", "no", "-soft_masking", "false",
            "-max_target_seqs", str(cfg["blastn_max_targets"]),
            "-num_threads", str(max(1, int(threads))),
            "-outfmt", BLASTN_OUTFMT,
            "-out", str(tab_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True)
        if proc.returncode != 0:
            raise RuntimeError(proc.stderr.strip() or "blastn failed")
        return parse_blastn_tabular(tab_path)

def _count_bracketed(tir: TirPair, intervals: List[Tuple[int, int]]) -> int:
    return sum(1 for s, e in intervals if tir.left_start <= s and e <= tir.right_end)

def _dinucleotide_entropy(seq: str) -> float:
    """Shannon entropy (bits) of overlapping ACGT dinucleotides."""
    if len(seq) < 2:
        return 0.0
    counts: Counter = Counter()
    seq = seq.upper()
    for i in range(len(seq) - 1):
        di = seq[i:i + 2]
        if "N" in di:
            continue
        counts[di] += 1
    total = sum(counts.values())
    if total == 0:
        return 0.0
    entropy = 0.0
    for c in counts.values():
        p = c / total
        entropy -= p * math.log2(p)
    return entropy

def _max_kmer_frac(seq: str, k: int) -> float:
    """Maximum fraction of any single k-mer across all phase offsets."""
    seq = seq.upper().encode("ascii", errors="ignore")
    n = len(seq)
    if n < k:
        return 0.0
    best = 0.0
    for phase in range(k):
        tiles = [seq[i:i + k] for i in range(phase, n - k + 1, k)]
        valid = [t for t in tiles if b"N" not in t]
        if not valid:
            continue
        counts = Counter(valid)
        best = max(best, max(counts.values()) / len(valid))
    return best

def _max_tandem_period_fraction(seq: str, max_period: int) -> float:
    """Maximum fraction of positions i where seq[i] == seq[i+p], p=1..max_period."""
    seq = seq.upper()
    n = len(seq)
    if n < 2 or max_period < 1:
        return 0.0
    best = 0.0
    for p in range(1, min(max_period, n - 1) + 1):
        total = 0
        matches = 0
        for i in range(n - p):
            a, b = seq[i], seq[i + p]
            if a == "N" or b == "N":
                continue
            total += 1
            if a == b:
                matches += 1
        if total:
            best = max(best, matches / total)
    return best

def _tir_is_low_complexity(seq: str, min_entropy: float, max_kmer_frac: float,
                           max_tandem_frac: float, max_tandem_period: int) -> Tuple[bool, str]:
    if not seq:
        return True, "empty TIR sequence"
    entropy = _dinucleotide_entropy(seq)
    if entropy < min_entropy:
        return True, f"dinuc_entropy={entropy:.2f}<{min_entropy}"
    for k in (1, 2, 3, 4):
        frac = _max_kmer_frac(seq, k)
        if frac > max_kmer_frac:
            return True, f"{k}-mer_fraction={frac:.2f}>{max_kmer_frac}"
    tandem_frac = _max_tandem_period_fraction(seq, max_tandem_period)
    if tandem_frac > max_tandem_frac:
        return True, f"tandem_fraction={tandem_frac:.2f}>{max_tandem_frac}"
    return False, ""

def _tir_brackets_reviewed_boundary(tir: TirPair, candidate_start: int, candidate_end: int, cfg: dict) -> bool:
    """Review-mode safety check: accepted TIR must reasonably bracket reviewed interval."""
    span = max(1, candidate_end - candidate_start + 1)
    slop = max(int(cfg.get("tir_edge_slop", 50_000)), int(span * float(cfg.get("tir_bracket_slop_frac", 0.10))))
    return tir.left_start <= candidate_start + slop and tir.right_end >= candidate_end - slop

def select_best_tir(pairs: List[TirPair], region_offset: int, candidate_start: int,
                    candidate_end: int, hallmark_intervals: List[Tuple[int, int]],
                    cfg: dict, region_seq: Optional[str] = None) -> Optional[TirPair]:
    valid = []
    min_entropy = cfg.get("tir_min_dinuc_entropy", 2.0)
    max_kmer_frac = cfg.get("tir_max_kmer_fraction", 0.70)
    max_tandem_frac = cfg.get("tir_max_tandem_fraction", 0.70)
    max_tandem_period = cfg.get("tir_tandem_max_period", 12)
    bracket_frac = cfg.get("tir_bracket_fraction", 1.0)
    for t in pairs:
        abs_t = TirPair(
            left_start=t.left_start + region_offset - 1,
            left_end=t.left_end + region_offset - 1,
            right_start=t.right_start + region_offset - 1,
            right_end=t.right_end + region_offset - 1,
            tir_length=t.tir_length,
            insert_size=t.insert_size,
            tir_identity=t.tir_identity,
            score=t.score,
            matches=t.matches,
            total=t.total,
            gaps=t.gaps,
            tir_evalue=t.tir_evalue,
        )
        if not (cfg["tir_min_len"] <= abs_t.tir_length <= cfg["tir_max_len"]):
            continue
        if abs_t.tir_identity < cfg["tir_min_id"]:
            continue
        if not (cfg["tir_min_insert"] <= abs_t.insert_size <= cfg["tir_max_insert"]):
            continue
        if region_seq is not None:
            left_seq = region_seq[t.left_start - 1:t.left_end]
            right_seq = region_seq[t.right_start - 1:t.right_end]
            left_lc, _ = _tir_is_low_complexity(left_seq, min_entropy, max_kmer_frac, max_tandem_frac, max_tandem_period)
            right_lc, _ = _tir_is_low_complexity(right_seq, min_entropy, max_kmer_frac, max_tandem_frac, max_tandem_period)
            if left_lc or right_lc:
                continue
        if not _tir_brackets_reviewed_boundary(abs_t, candidate_start, candidate_end, cfg):
            continue
        edge_distance = abs(abs_t.left_start - candidate_start) + abs(candidate_end - abs_t.right_end)
        bracketed = _count_bracketed(abs_t, hallmark_intervals)
        if hallmark_intervals and bracketed < max(1, math.ceil(len(hallmark_intervals) * bracket_frac)):
            continue
        valid.append((bracketed, -edge_distance, abs_t.tir_identity, abs_t.insert_size, abs_t.tir_length, abs_t.score, abs_t))
    if not valid:
        return None

    valid.sort(key=lambda x: x[:-1], reverse=True)
    return valid[0][-1]

def find_tsd(left_flank: str, right_flank: str, k_min: int, k_max: int, max_slide: int) -> Optional[Tsd]:
    left = left_flank.upper()
    right = right_flank.upper()
    best: Optional[Tsd] = None
    for k in range(k_max, k_min - 1, -1):
        if k > len(left) or k > len(right):
            continue
        max_mm = 0 if k <= 5 else (1 if k <= 8 else 2)
        for sl in range(max_slide + 1):
            if k + sl > len(left):
                break
            lk = left[len(left) - k - sl: len(left) - sl] if sl > 0 else left[-k:]
            for sr in range(max_slide + 1):
                if k + sr > len(right):
                    break
                rk = right[sr:sr + k]
                if "N" in lk or "N" in rk:
                    continue
                mm = sum(1 for a, b in zip(lk, rk) if a != b)
                if mm <= max_mm:
                    cand = Tsd(lk, rk, k, mm, 100.0 * (k - mm) / k, sl, sr)
                    if best is None or (cand.length, cand.identity) > (best.length, best.identity):
                        best = cand
        if best is not None and best.length == k:
            return best
    return best


def hallmark_intervals(marker: pd.DataFrame, geve_name: str, start: int, end: int) -> List[Tuple[int, int]]:
    if marker.empty:
        return []
    q = marker[(marker["geve_name"].astype(str) == geve_name) & (marker["feature"].astype(str) == "hallmark")]
    out = []
    for _, r in q.iterrows():
        s, e = _safe_int(r.get("start")), _safe_int(r.get("end"))
        if s is not None and e is not None and s >= start and e <= end:
            out.append((s, e))
    return out

def redetect_tir_tsd(seqs: Dict[str, str], marker: pd.DataFrame, geve_name: str,
                     contig: str, start: int, end: int, cfg: dict,
                     threads: int) -> Tuple[Optional[TirPair], Optional[Tsd], List[str]]:
    messages: List[str] = []
    if not seqs:
        messages.append("genome FASTA not provided; TIR/TSD detection skipped")
        return None, None, messages
    contig_seq = seqs.get(str(contig), "")
    if not contig_seq:
        messages.append(f"sequence unavailable for {contig}; TIR/TSD detection skipped")
        return None, None, messages
    search_start = max(1, start - cfg["tir_flank"])
    search_end = min(len(contig_seq), end + cfg["tir_flank"])
    region_seq = fetch_seq(seqs, contig, search_start, search_end)
    if not region_seq:
        messages.append(f"sequence unavailable for {contig}:{search_start}-{search_end}; TIR/TSD detection skipped")
        return None, None, messages
    try:
        raw_pairs = run_blastn_self(region_seq, cfg, threads=threads)
    except Exception as exc:
        messages.append(f"TIR detection skipped/failed: {exc}")
        return None, None, messages
    hms = hallmark_intervals(marker, geve_name, start, end)
    tir = select_best_tir(raw_pairs, search_start, start, end, hms, cfg, region_seq=region_seq)
    if tir is None:
        messages.append(f"no TIR passed filters among {len(raw_pairs)} raw inverted-repeat pairs")
        return None, None, messages
    left_flank = fetch_seq(seqs, contig, max(1, tir.left_start - cfg["tsd_flank"]), tir.left_start - 1)
    right_flank = fetch_seq(seqs, contig, tir.right_end + 1, tir.right_end + cfg["tsd_flank"])
    tsd = find_tsd(left_flank, right_flank, cfg["tsd_min_len"], cfg["tsd_max_len"], cfg["tsd_max_slide"])
    messages.append(
        f"TIR detected: {tir.left_start}-{tir.left_end} / {tir.right_start}-{tir.right_end}, "
        f"identity={tir.tir_identity:.2f}%"
    )
    if tsd is None:
        messages.append("TSD not detected")
    else:
        messages.append(f"TSD detected: {tsd.sequence_left}|{tsd.sequence_right}, len={tsd.length}")
    return tir, tsd, messages

def get_original_tir_tsd(marker: pd.DataFrame, geve_name: str) -> Tuple[Optional[TirPair], Optional[Tsd]]:
    q = marker[marker["geve_name"].astype(str) == geve_name]
    left = q[q["feature"].astype(str) == "TIR_left"]
    right = q[q["feature"].astype(str) == "TIR_right"]
    tir = None
    if not left.empty and not right.empty:
        l, r = left.iloc[0], right.iloc[0]
        ls, le = _safe_int(l.get("start")), _safe_int(l.get("end"))
        rs, re_ = _safe_int(r.get("start")), _safe_int(r.get("end"))
        if None not in (ls, le, rs, re_):
            score = _safe_int(l.get("score"), 0) or 0
            length = le - ls + 1
            tir = TirPair(ls, le, rs, re_, length, rs - le - 1, float("nan"), score, 0, 0, 0)
    tsd = None
    t5 = q[q["feature"].astype(str) == "TSD_5p"]
    t3 = q[q["feature"].astype(str) == "TSD_3p"]
    if not t5.empty and not t3.empty:
        a, b = t5.iloc[0], t3.iloc[0]
        sleft = str(a.get("name", "") or "")
        sright = str(b.get("name", "") or "")
        length = len(sleft) if sleft and sleft != "." else len(sright)
        ident = _safe_float(a.get("score"), 100.0)
        mismatches = int(round(length * (100.0 - ident) / 100.0)) if length else 0
        tsd = Tsd(sleft, sright, length, mismatches, ident, 0, 0)
    return tir, tsd

def tir_fields(tir: Optional[TirPair]) -> dict:
    if tir is None:
        return dict(tir_length="NA", tir_score="NA", tir_identity_pct="NA", tir_gaps="NA")
    return dict(
        tir_length=tir.tir_length,
        tir_score=tir.score,
        tir_identity_pct="NA" if math.isnan(tir.tir_identity) else round(tir.tir_identity, 2),
        tir_gaps=tir.gaps,
    )

def tsd_fields(tsd: Optional[Tsd]) -> dict:
    if tsd is None:
        return dict(tsd_len="NA", tsd_left="NA", tsd_right="NA", tsd_mismatch="NA", tsd_conservation="NODETECT")
    return dict(
        tsd_len=tsd.length,
        tsd_left=tsd.sequence_left,
        tsd_right=tsd.sequence_right,
        tsd_mismatch=tsd.mismatches,
        tsd_conservation="PERFECT" if tsd.mismatches == 0 else "IMPERFECT",
    )

def build_reviewed_records(review: pd.DataFrame, summary: pd.DataFrame, marker: pd.DataFrame,
                           seqs: Dict[str, str], cfg: dict, threads: int) -> Tuple[List[dict], List[str]]:
    summary_by_name = {str(r["geve_name"]): r for _, r in summary.iterrows()}
    records: List[dict] = []
    messages: List[str] = []
    kept = review[review["action"] != "remove"].copy().reset_index(drop=True)
    for idx, row in kept.iterrows():
        old_name = row["geve_name"]
        new_name = re.sub(r"_GEVE_\d+$", f"_GEVE_{idx + 1:03d}", old_name)
        if new_name == old_name:
            prefix = old_name.split("_GEVE_", 1)[0] if "_GEVE_" in old_name else infer_prefix(Path("findGEVE.summary.tsv"))
            new_name = f"{prefix}_GEVE_{idx + 1:03d}"
        original = summary_by_name[old_name]
        contig = row["contig"]
        action = row["action"]
        if action == "change":
            candidate_start = int(row["review_start"])
            candidate_end = int(row["review_end"])
            tir, tsd, tir_messages = redetect_tir_tsd(seqs, marker, old_name, contig, candidate_start, candidate_end, cfg, threads)
            messages.extend([f"{old_name}: {m}" for m in tir_messages])
            final_start = tir.left_start if tir is not None else candidate_start
            final_end = tir.right_end if tir is not None else candidate_end
            boundary_method = "reviewed_tir_boundary" if tir is not None else "reviewed_manual_boundary"
        else:
            candidate_start = int(row["original_start"])
            candidate_end = int(row["original_end"])
            final_start = candidate_start
            final_end = candidate_end
            tir, tsd = get_original_tir_tsd(marker, old_name)
            boundary_method = "original_boundary"
        geve_len = final_end - final_start + 1
        seq = fetch_seq(seqs, contig, final_start, final_end)
        gc = gc_of_seq(seq) if seq else _safe_float(original.get("gc"))
        records.append(dict(
            original_geve_name=old_name,
            reviewed_geve_name=new_name,
            action=action,
            contig=contig,
            candidate_start=candidate_start,
            candidate_end=candidate_end,
            geve_start=final_start,
            geve_end=final_end,
            geve_length=geve_len,
            gc_geve=gc,
            tir=tir,
            tsd=tsd,
            has_tir=tir is not None,
            boundary_method=boundary_method,
            original_summary=original.to_dict(),
        ))
    return records, messages

def write_reviewed_summary(records: List[dict], path: Path) -> None:
    rows = []
    for r in records:
        old = dict(r["original_summary"])
        row = dict(old)
        row["original_geve_name"] = r["original_geve_name"]
        row["geve_name"] = r["reviewed_geve_name"]
        row["review_action"] = r["action"]
        row["boundary_method"] = r["boundary_method"]
        row["contig_id"] = r["contig"]
        row["start"] = r["geve_start"]
        row["end"] = r["geve_end"]
        row["geve_length"] = r["geve_length"]
        row["gc"] = "NA" if math.isnan(r["gc_geve"]) else round(r["gc_geve"], 2)
        row["has_tir"] = "yes" if r["has_tir"] else "no"
        row.update(tir_fields(r["tir"]))
        row.update(tsd_fields(r["tsd"]))
        rows.append(row)
    pd.DataFrame(rows).to_csv(path, sep="\t", index=False)

def write_reviewed_fna(records: List[dict], seqs: Dict[str, str], path: Path) -> None:
    with path.open("w") as fh:
        for r in records:
            seq = fetch_seq(seqs, r["contig"], r["geve_start"], r["geve_end"])
            header = (
                f">{r['reviewed_geve_name']} contig={r['contig']} start={r['geve_start']} "
                f"end={r['geve_end']} length={r['geve_length']} boundary_method={r['boundary_method']}"
            )
            fh.write(header + "\n")
            fh.write(wrap_fasta(seq) + "\n")

def feature_rows_for_record(marker: pd.DataFrame, record: dict, include_flank: bool = False) -> pd.DataFrame:
    old = record["original_geve_name"]
    start = record["geve_start"]
    end = record["geve_end"]
    if include_flank:
        flank = viz_flank_size(record["geve_length"])
        start = max(1, start - flank)
        end = end + flank
    q = marker[(marker["geve_name"].astype(str) == old) & (~marker["feature"].astype(str).isin(FEATURE_IGNORE))].copy()
    if q.empty:
        return q
    q["start_i"] = q["start"].map(_safe_int)
    q["end_i"] = q["end"].map(_safe_int)
    q = q.dropna(subset=["start_i", "end_i"])
    q["start_i"] = q["start_i"].astype(int)
    q["end_i"] = q["end_i"].astype(int)
    q = q[(q["end_i"] >= start) & (q["start_i"] <= end)].copy()
    return q.sort_values(["start_i", "end_i"], kind="mergesort")

def write_reviewed_cds_pep(records: List[dict], marker: pd.DataFrame, seqs: Dict[str, str], cds_path: Path,
                           pep_path: Path, hallmark_dir: Path, prefix: str) -> None:
    hallmark_best: Dict[str, Dict[str, Tuple[str, str]]] = {}
    with cds_path.open("w") as cds_fh, pep_path.open("w") as pep_fh:
        for r in records:
            gid = r["reviewed_geve_name"]
            feats = feature_rows_for_record(marker, r, include_flank=False)
            for orf_idx, (_, row) in enumerate(feats.iterrows(), start=1):
                s, e = int(row["start_i"]), int(row["end_i"])
                strand = str(row.get("strand", "+") or "+")
                cds = fetch_seq(seqs, r["contig"], s, e)
                if strand == "-":
                    cds = revcomp(cds)
                pep = translate_cds(cds)
                feature = str(row.get("feature", "orf") or "orf")
                name = str(row.get("name", ".") or ".")
                label = f"orf{orf_idx:05d}"
                annot = name if feature == "hallmark" and name not in {"", "."} else ""
                cds_header = f">{gid}_{label}"
                pep_header = f">{gid}_{label}"
                if annot:
                    cds_header += f" {annot}"
                    pep_header += f" {annot}"
                cds_header += f" contig={r['contig']} start={s} end={e} strand={strand} length={len(cds)}"
                pep_header += f" length={len(pep)}"
                cds_fh.write(cds_header + "\n" + wrap_fasta(cds) + "\n")
                pep_fh.write(pep_header + "\n" + wrap_fasta(pep) + "\n")
                if feature == "hallmark" and annot:
                    current = hallmark_best.setdefault(annot, {}).get(gid)
                    if current is None or len(pep) > len(current[1]):
                        hallmark_best[annot][gid] = (f">{gid}_{annot}", pep)
    hallmark_dir.mkdir(parents=True, exist_ok=True)
    for hallmark, geve_map in sorted(hallmark_best.items(), key=lambda kv: _natural_key(kv[0])):
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", hallmark.lower())
        out = hallmark_dir / f"{prefix}.{safe}.pep"
        with out.open("w") as fh:
            for gid, (header, pep) in sorted(geve_map.items(), key=lambda kv: _natural_key(kv[0])):
                fh.write(header + "\n" + wrap_fasta(pep) + "\n")

def write_reviewed_gff3(records: List[dict], marker: pd.DataFrame, path: Path) -> None:
    with path.open("w") as fh:
        fh.write("##gff-version 3\n")
        for r in records:
            gid = r["reviewed_geve_name"]
            contig = r["contig"]
            fh.write(
                f"{contig}\tfindGEVE_review\tmobile_genetic_element\t{r['geve_start']}\t{r['geve_end']}\t.\t+\t.\t"
                f"ID={gid};Name={gid};original_geve_name={r['original_geve_name']};boundary_method={r['boundary_method']}\n"
            )
            feats = feature_rows_for_record(marker, r, include_flank=False)
            for orf_idx, (_, row) in enumerate(feats.iterrows(), start=1):
                feature = str(row.get("feature", "ORF") or "ORF")
                name = str(row.get("name", ".") or ".")
                strand = str(row.get("strand", ".") or ".")
                score = str(row.get("score", ".") or ".")
                fh.write(
                    f"{contig}\tfindGEVE_review\t{feature}\t{int(row['start_i'])}\t{int(row['end_i'])}\t{score}\t{strand}\t.\t"
                    f"ID={gid}.orf{orf_idx:05d};Parent={gid};Name={name}\n"
                )

def write_reviewed_markerout(records: List[dict], marker: pd.DataFrame, seqs: Dict[str, str], path: Path) -> None:
    with path.open("w") as fh:
        fh.write("contig\tgeve_name\tfeature\tname\tstart\tend\tstrand\te_value\tscore\n")
        for r in records:
            gid = r["reviewed_geve_name"]
            old = r["original_geve_name"]
            contig = r["contig"]
            gstart, gend = r["geve_start"], r["geve_end"]
            clen = len(seqs.get(contig, "")) if seqs else gend
            flank = viz_flank_size(r["geve_length"])
            region_start = max(1, gstart - flank)
            region_end = min(clen, gend + flank) if clen else gend + flank
            fh.write(f"{contig}\t{gid}\tGEVE\t.\t{gstart}\t{gend}\t.\tNA\t{r['geve_length']}\n")
            if region_start < gstart:
                fh.write(f"{contig}\t{gid}\tflank_left\t.\t{region_start}\t{gstart - 1}\t.\tNA\tNA\n")
            if region_end > gend:
                fh.write(f"{contig}\t{gid}\tflank_right\t.\t{gend + 1}\t{region_end}\t.\tNA\tNA\n")
            tir = r["tir"]
            if tir is not None:
                fh.write(f"{contig}\t{gid}\tTIR_left\t.\t{tir.left_start}\t{tir.left_end}\t+\tNA\t{tir.score}\n")
                fh.write(f"{contig}\t{gid}\tTIR_right\t.\t{tir.right_start}\t{tir.right_end}\t-\tNA\t{tir.score}\n")
            tsd = r["tsd"]
            if tsd is not None and tir is not None:
                ltsd_end = tir.left_start - 1 - tsd.left_shift
                ltsd_start = ltsd_end - tsd.length + 1
                rtsd_start = tir.right_end + 1 + tsd.right_shift
                rtsd_end = rtsd_start + tsd.length - 1
                fh.write(f"{contig}\t{gid}\tTSD_5p\t{tsd.sequence_left}\t{ltsd_start}\t{ltsd_end}\t+\tNA\t{tsd.identity:.1f}\n")
                fh.write(f"{contig}\t{gid}\tTSD_3p\t{tsd.sequence_right}\t{rtsd_start}\t{rtsd_end}\t+\tNA\t{tsd.identity:.1f}\n")
            q = marker[(marker["geve_name"].astype(str) == old) & (~marker["feature"].astype(str).isin(FEATURE_IGNORE))].copy()
            for _, row in q.iterrows():
                s, e = _safe_int(row.get("start")), _safe_int(row.get("end"))
                if s is None or e is None or e < region_start or s > region_end:
                    continue
                vals = [
                    contig, gid, row.get("feature", "."), row.get("name", "."),
                    str(s), str(e), row.get("strand", "."), row.get("e_value", "NA"), row.get("score", "NA"),
                ]
                fh.write("\t".join(map(str, vals)) + "\n")

def write_reviewed_bed(records: List[dict], bed: pd.DataFrame, path: Path) -> None:
    if bed.empty:
        path.write_text("contig_id\twindow_start\twindow_end\tgeve_name\trel_start\trel_end\tregion_type\tgc\trolling_score_mean\tn_orfs\tgvog_hits\tpfam_hits\n")
        return
    out_rows = []
    for r in records:
        gid = r["reviewed_geve_name"]
        old = r["original_geve_name"]
        gstart, gend = r["geve_start"], r["geve_end"]
        flank = viz_flank_size(r["geve_length"])
        region_start = max(1, gstart - flank)
        region_end = gend + flank
        q = bed[bed["geve_name"].astype(str) == old].copy()
        if q.empty:
            continue
        for col in ["window_start", "window_end"]:
            q[col] = pd.to_numeric(q[col], errors="coerce")
        q = q[(q["window_end"] >= region_start) & (q["window_start"] <= region_end)].copy()
        if q.empty:
            continue
        centers = ((q["window_start"] + q["window_end"]) / 2.0)
        q["geve_name"] = gid
        q["region_type"] = ["flank_left" if c < gstart else ("geve" if c <= gend else "flank_right") for c in centers]
        out_rows.append(q)
    if out_rows:
        pd.concat(out_rows, ignore_index=True).to_csv(path, sep="\t", index=False)
    else:
        path.write_text("contig_id\twindow_start\twindow_end\tgeve_name\trel_start\trel_end\tregion_type\tgc\trolling_score_mean\tn_orfs\tgvog_hits\tpfam_hits\n")


def run_geve_plot(marker_path: Path, bed_path: Path, outdir: Path) -> None:
    candidates = [
        Path(__file__).resolve().parent / "findGEVE_plot_v5.py",
        Path(__file__).resolve().parent / "findGEVE_plot.py",
    ]
    script = next((p for p in candidates if p.is_file()), None)
    if script is None:
        _LOG.warning("findGEVE_plot.py not found; skipping plot step")
        return
    if not bed_path.is_file() or not bed_path.read_text().strip():
        _LOG.warning("reviewed geve.bed is empty; skipping plot step")
        return
    cmd = [sys.executable, str(script), str(marker_path.resolve()), str(bed_path.resolve())]
    _LOG.info(f"Plotting GEVEs: {' '.join(cmd)}")
    try:
        proc = subprocess.run(cmd, cwd=str(outdir), stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
    except OSError as exc:
        _LOG.warning(f"Failed to launch findGEVE_plot.py: {exc}")
        return
    for line in proc.stdout.splitlines():
        if line.strip():
            _LOG.info(f"[PLOT] {line}")
    for line in proc.stderr.splitlines():
        if line.strip():
            _LOG.warning(f"[PLOT] {line}")
    if proc.returncode != 0:
        _LOG.warning(f"findGEVE_plot.py exited with code {proc.returncode}; no plot produced")

from collections import defaultdict
from urllib.parse import unquote
import numpy as np
import pyhmmer
import pyrodigal_gv

_HALLMARK_ORDER = ["A32", "D5", "SFII", "MCP", "mRNAc", "PolB", "RNAPL", "RNAPS", "RNR", "VLTF3"]
_HALLMARK_CANONICAL = {x.lower(): x for x in _HALLMARK_ORDER}
_HALLMARK_SCORE_CUTOFFS = {"A32":100.0,"D5":180.0,"SFII":120.0,"MCP":120.0,"mRNAc":180.0,"PolB":300.0,"RNAPL":300.0,"RNAPS":250.0,"RNR":200.0,"VLTF3":100.0}
_HOST_ORF_EDGE_OVERLAP_BP = 15
_HOST_ORF_EDGE_OVERLAP_FRACTION = 0.02

@dataclass
class ReviewOrf:
    orf_id: str
    contig: str
    start: int
    end: int
    strand: int
    protein: str
    hallmark: Optional[str] = None
    hallmark_bitscore: float = 0.0
    hallmark_evalue: float = float("inf")
    gvog: Optional[str] = None
    gvog_bitscore: float = 0.0
    gvog_evalue: float = float("inf")
    best_pfam_acc: Optional[str] = None
    best_pfam_name: Optional[str] = None
    best_pfam_bitscore: float = 0.0
    best_pfam_evalue: float = float("inf")
    virbit: float = 0.0
    pfambit: float = 0.0
    net_score: float = 0.0

def _canon_hallmark(name: str) -> str:
    return _HALLMARK_CANONICAL.get(str(name).lower(), str(name))

def _hallmark_key(name: str):
    c = _canon_hallmark(name)
    return (0, _HALLMARK_ORDER.index(c)) if c in _HALLMARK_ORDER else (1, _natural_key(c))

def _parse_attrs(text: str) -> Dict[str, str]:
    out = {}
    for raw in str(text or "").split(";"):
        part = raw.strip()
        if not part:
            continue
        if "=" in part:
            k, v = part.split("=", 1)
        elif " " in part:
            k, v = part.split(None, 1)
            v = v.strip().strip('"')
        else:
            continue
        out[k.strip()] = unquote(v.strip().strip('"'))
    return out

def _merge_host_intervals(intervals: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    merged = []
    for s, e in sorted((min(s,e), max(s,e)) for s,e in intervals):
        if merged and s <= merged[-1][1] + 1:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
        else:
            merged.append((s,e))
    return merged

def parse_host_gff(path: Optional[Path]) -> Dict[str, List[Tuple[int, int]]]:
    """Parse the host annotation exactly as findGEVE does."""
    if path is None:
        return {}
    if not path.is_file():
        raise SystemExit(f"Error: GFF file not found: {path}")

    gene_like = {
        "gene", "mrna", "transcript", "primary_transcript",
        "lncrna", "ncrna", "rrna", "trna", "snorna", "snrna", "mirna",
        "pseudogene", "pseudogenic_transcript",
    }
    parts = {"cds", "exon"}
    ignored = {
        "intron", "start_codon", "stop_codon", "five_prime_utr",
        "three_prime_utr", "5utr", "3utr", "utr",
    }
    by_contig = defaultdict(list)
    by_parent = defaultdict(list)
    n_rows = n_used = n_gene_like = n_part = 0

    opener = gzip.open if str(path).endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            c = line.split("\t")
            if len(c) < 5:
                continue
            n_rows += 1
            contig = c[0]
            feature = c[2].strip().lower() if len(c) > 2 else ""
            try:
                start = int(c[3])
                end = int(c[4])
            except (TypeError, ValueError):
                continue
            if start <= 0 or end <= 0:
                continue
            start, end = min(start, end), max(start, end)

            if feature in ignored:
                continue
            attrs = _parse_attrs(c[8] if len(c) >= 9 else "")

            if feature in gene_like:
                by_contig[contig].append((start, end))
                n_used += 1
                n_gene_like += 1
                continue

            if feature in parts:
                by_contig[contig].append((start, end))
                n_used += 1
                n_part += 1
                parent_text = attrs.get("Parent") or attrs.get("parent")
                if parent_text:
                    parents = [x.strip() for x in parent_text.split(",") if x.strip()]
                else:
                    fallback = (
                        attrs.get("transcript_id") or attrs.get("gene_id")
                        or attrs.get("ID") or attrs.get("Name")
                    )
                    parents = [fallback] if fallback else []
                for parent in parents:
                    by_parent[(contig, parent)].append((start, end))
                continue

            if any(tok in feature for tok in ("gene", "transcript", "mrna")):
                by_contig[contig].append((start, end))
                n_used += 1
                n_gene_like += 1
            elif any(tok == feature or tok in feature for tok in ("cds", "exon")):
                by_contig[contig].append((start, end))
                n_used += 1
                n_part += 1

    for (contig, _parent), spans in by_parent.items():
        if spans:
            by_contig[contig].append((
                min(start for start, _ in spans),
                max(end for _, end in spans),
            ))

    out = {contig: _merge_host_intervals(v) for contig, v in by_contig.items() if v}
    n_intervals = sum(len(v) for v in out.values())
    _LOG.info(
        f"Host GFF mask: parsed {n_rows:,} feature row(s) from {path}; "
        f"used {n_used:,} row(s) ({n_gene_like:,} gene/transcript-like, "
        f"{n_part:,} CDS/exon); collapsed to {n_intervals:,} interval(s) "
        f"on {len(out):,} contig(s)"
    )
    if n_intervals == 0:
        _LOG.warning(
            "Host GFF mask: no usable gene/mRNA/transcript/CDS/exon intervals "
            "were parsed; no ORFs will be removed by --gff"
        )
    return out

def _host_overlap_bp(
    intervals: List[Tuple[int, int]],
    start: int,
    end: int,
) -> int:
    if not intervals or end < start:
        return 0

    lo, hi = 0, len(intervals)
    while lo < hi:
        mid = (lo + hi) // 2
        if intervals[mid][1] < start:
            lo = mid + 1
        else:
            hi = mid

    overlap = 0
    i = lo
    while i < len(intervals):
        iv_start, iv_end = intervals[i]
        if iv_start > end:
            break
        overlap += max(0, min(end, iv_end) - max(start, iv_start) + 1)
        i += 1
    return overlap

def _host_overlap_allowed(
    orf: ReviewOrf,
    intervals: List[Tuple[int, int]],
    overlap_bp: int,
) -> bool:
    if overlap_bp <= 0:
        return True

    orf_len = max(1, orf.end - orf.start + 1)
    tolerance = max(
        _HOST_ORF_EDGE_OVERLAP_BP,
        int(np.ceil(orf_len * _HOST_ORF_EDGE_OVERLAP_FRACTION)),
    )
    if overlap_bp > tolerance:
        return False

    overlaps: List[Tuple[int, int]] = []
    for iv_start, iv_end in intervals:
        if iv_end < orf.start:
            continue
        if iv_start > orf.end:
            break
        s = max(orf.start, iv_start)
        e = min(orf.end, iv_end)
        if s <= e:
            overlaps.append((s, e))
            if len(overlaps) > 1:
                return False

    if len(overlaps) != 1:
        return False
    s, e = overlaps[0]
    return s == orf.start or e == orf.end

def predict_contig_orfs(
    contig: str,
    contig_seq: str,
    host: Dict[str, List[Tuple[int, int]]],
) -> List[ReviewOrf]:
    """Run Pyrodigal-GV on the complete contig, matching findGEVE stage 1.

    Predicting independently on a cropped GEVE interval changes the sequence
    context at both ends and can split, merge, or remove ORFs. Whole-contig
    prediction makes boundary extension monotonic with respect to the fixed
    contig ORF catalogue: a larger interval can only retain the same or more
    fully contained ORFs.
    """
    try:
        genes = pyrodigal_gv.ViralGeneFinder(meta=True).find_genes(
            contig_seq.encode("ascii")
        )
    except Exception as exc:
        raise SystemExit(
            f"Error: pyrodigal-gv failed on complete contig {contig}: {exc}"
        ) from exc

    out: List[ReviewOrf] = []
    masked = 0
    edge_only = 0
    predicted = 0
    intervals = host.get(contig, [])
    for i, gene in enumerate(genes, 1):
        predicted += 1
        start = int(gene.begin)
        end = int(gene.end)
        protein = gene.translate().rstrip("*")
        if not protein:
            continue
        orf = ReviewOrf(
            f"{contig}__orf{i:05d}", contig, start, end,
            int(gene.strand), protein,
        )
        overlap_bp = _host_overlap_bp(intervals, start, end)
        if overlap_bp > 0:
            if _host_overlap_allowed(orf, intervals, overlap_bp):
                edge_only += 1
            else:
                masked += 1
                continue
        out.append(orf)

    _LOG.info(
        f"Pyrodigal-GV whole-contig prediction | {contig}: "
        f"predicted={predicted:,}, host-masked={masked:,}, "
        f"edge-overlap-retained={edge_only:,}, retained={len(out):,}"
    )
    return out


def select_orfs_in_record(record: dict, contig_orfs: List[ReviewOrf]) -> List[ReviewOrf]:
    """Use the same fully-contained ORF rule as findGEVE."""
    start = int(record["geve_start"])
    end = int(record["geve_end"])
    return [o for o in contig_orfs if o.start >= start and o.end <= end]

def _digital(orfs: List[ReviewOrf]):
    aa = pyhmmer.easel.Alphabet.amino()
    return [pyhmmer.easel.TextSequence(name=o.orf_id.encode(), sequence=o.protein).digitize(aa) for o in orfs if len(o.protein) <= 100000]

def _hmm_name(hits) -> str:
    try: x = hits.query.name
    except AttributeError: x = hits.query_name
    return x.decode() if isinstance(x, bytes) else str(x)

def scan_changed_orfs(orfs: List[ReviewOrf], db: Path, evalue: float, threads: int) -> None:
    """Annotate whole-contig ORFs using the same scan scope as findGEVE."""
    if not orfs:
        return
    hallmark_db = db / "NCLDV_markers.hmm"
    gvog_db = db / "gvog.complete.hmm"
    pfam_db = db / "Pfam-A.hmm"
    for req in (hallmark_db, gvog_db):
        if not req.is_file():
            raise SystemExit(f"Error: required database file not found: {req}")

    by_id = {o.orf_id: o for o in orfs}
    hallmark_seqs = _digital(orfs)
    if not hallmark_seqs:
        _LOG.warning("No reviewed ORFs were eligible for HMM scanning")
        return

    hallmark_positive_contigs = set()
    with pyhmmer.plan7.HMMFile(str(hallmark_db)) as hf:
        hmms = list(hf)
    for hits in pyhmmer.hmmsearch(
        hmms, hallmark_seqs, cpus=max(1, threads), E=evalue
    ):
        name = _canon_hallmark(_hmm_name(hits))
        cutoff = _HALLMARK_SCORE_CUTOFFS.get(name, 0.0)
        for hit in hits:
            if not hit.included or float(hit.score) < cutoff:
                continue
            target = hit.name.decode() if isinstance(hit.name, bytes) else str(hit.name)
            o = by_id.get(target)
            if o is None:
                continue
            score = float(hit.score)
            if score > o.hallmark_bitscore:
                o.hallmark = name
                o.hallmark_bitscore = score
                o.hallmark_evalue = float(hit.evalue)
            if score > o.virbit:
                o.virbit = score
            hallmark_positive_contigs.add(o.contig)

    # findGEVE scans GVOG/Pfam only on hallmark-positive contigs.
    downstream_orfs = [o for o in orfs if o.contig in hallmark_positive_contigs]
    downstream_seqs = _digital(downstream_orfs)
    downstream_by_id = {o.orf_id: o for o in downstream_orfs}

    if downstream_seqs:
        with pyhmmer.plan7.HMMFile(str(gvog_db)) as hf:
            for hits in pyhmmer.hmmsearch(
                hf, downstream_seqs, cpus=max(1, threads), E=evalue
            ):
                name = _hmm_name(hits)
                for hit in hits:
                    if not hit.included:
                        continue
                    target = hit.name.decode() if isinstance(hit.name, bytes) else str(hit.name)
                    o = downstream_by_id.get(target)
                    if o is None:
                        continue
                    score = float(hit.score)
                    if score > o.gvog_bitscore:
                        o.gvog = name
                        o.gvog_bitscore = score
                        o.gvog_evalue = float(hit.evalue)
                    if score > o.virbit:
                        o.virbit = score

        if pfam_db.is_file():
            with pyhmmer.plan7.HMMFile(str(pfam_db)) as hf:
                for hits in pyhmmer.hmmsearch(
                    hf, downstream_seqs, cpus=max(1, threads), E=evalue
                ):
                    name = _hmm_name(hits)
                    try:
                        raw = hits.query.accession
                        acc = (raw.decode() if isinstance(raw, bytes) else raw) or name
                        acc = acc.split(".")[0]
                    except AttributeError:
                        acc = name
                    for hit in hits:
                        if not hit.included:
                            continue
                        target = hit.name.decode() if isinstance(hit.name, bytes) else str(hit.name)
                        o = downstream_by_id.get(target)
                        if o is None:
                            continue
                        score = float(hit.score)
                        if score > o.best_pfam_bitscore:
                            o.best_pfam_acc = acc
                            o.best_pfam_name = name
                            o.best_pfam_bitscore = score
                            o.best_pfam_evalue = float(hit.evalue)
                            o.pfambit = score

    for o in orfs:
        o.net_score = o.virbit - max(0.0, o.pfambit - o.virbit)

def read_peptide_fasta(path: Optional[Path]) -> Dict[str, str]:
    if path is None or not path.is_file():
        return {}
    peptides: Dict[str, List[str]] = {}
    current: Optional[str] = None
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        for raw in fh:
            line = raw.strip()
            if not line:
                continue
            if line.startswith(">"):
                current = line[1:].split()[0]
                peptides[current] = []
            elif current is not None:
                peptides[current].append(line)
    return {key: "".join(parts) for key, parts in peptides.items()}


def original_orfs(record: dict, marker: pd.DataFrame, seqs: Dict[str,str], original_peptides: Optional[Dict[str, str]] = None) -> List[ReviewOrf]:
    original_peptides = original_peptides or {}
    q = marker[(marker["geve_name"].astype(str) == record["original_geve_name"]) & (~marker["feature"].astype(str).isin(FEATURE_IGNORE))].copy()
    out = []
    q["_s"] = q["start"].map(_safe_int); q["_e"] = q["end"].map(_safe_int)
    q = q.dropna(subset=["_s","_e"]).sort_values(["_s","_e"])
    for i, (_, row) in enumerate(q.iterrows(), 1):
        s,e = int(row["_s"]), int(row["_e"])
        if s < record["geve_start"] or e > record["geve_end"]: continue
        strand = -1 if str(row.get("strand","+")) == "-" else 1
        cds = fetch_seq(seqs, record["contig"], s, e)
        if strand < 0: cds = revcomp(cds)
        original_id = f"{record['original_geve_name']}_orf{i:05d}"
        protein = original_peptides.get(original_id, translate_cds(cds))
        o = ReviewOrf(f"{record['reviewed_geve_name']}__orf{i:05d}", record["contig"], s,e,strand,protein)
        feature, name = str(row.get("feature","")), str(row.get("name","") or "")
        score, ev = _safe_float(row.get("score"),0.0), _safe_float(row.get("e_value"),float("inf"))
        if feature == "hallmark": o.hallmark, o.hallmark_bitscore, o.hallmark_evalue = _canon_hallmark(name), score, ev; o.virbit=score
        elif feature == "gvog": o.gvog, o.gvog_bitscore, o.gvog_evalue = name, score, ev; o.virbit=score
        elif feature == "pfam": o.best_pfam_name, o.best_pfam_acc, o.best_pfam_bitscore, o.best_pfam_evalue = name,name,score,ev; o.pfambit=score
        o.net_score = o.virbit - max(0.0, o.pfambit - o.virbit)
        out.append(o)
    return out

def _valid_original_tir(marker: pd.DataFrame, old: str, start: int, end: int):
    tir, tsd = get_original_tir_tsd(marker, old)
    if tir and tir.left_start >= start and tir.right_end <= end:
        return tir, tsd
    return None, None

def build_v2_records(
    review: pd.DataFrame,
    summary: pd.DataFrame,
    marker: pd.DataFrame,
    seqs: Dict[str, str],
    host: Dict[str, List[Tuple[int, int]]],
    db: Path,
    evalue: float,
    threads: int,
    prefix: str,
    original_peptides: Optional[Dict[str, str]] = None,
):
    smap = {str(r["geve_name"]): r for _, r in summary.iterrows()}
    temp = []
    for _, row in review[review["action"] != "remove"].iterrows():
        start = int(row["review_start"] if row["action"] == "change" else row["original_start"])
        end = int(row["review_end"] if row["action"] == "change" else row["original_end"])
        contig = str(row["contig"])
        if contig not in seqs:
            raise SystemExit(f"Error: contig not found in genome: {contig}")
        if start < 1 or end > len(seqs[contig]) or start >= end:
            raise SystemExit(
                f"Error: invalid reviewed coordinates for {row['geve_name']}: "
                f"{contig}:{start}-{end} (contig length {len(seqs[contig])})"
            )
        temp.append(dict(
            original_geve_name=row["geve_name"],
            action=row["action"],
            contig=contig,
            geve_start=start,
            geve_end=end,
            geve_length=end - start + 1,
            original_summary=smap[row["geve_name"]].to_dict(),
        ))

    for record in temp:
        record["reviewed_geve_name"] = record["original_geve_name"]
        record["geve_id"] = record["original_geve_name"]
        record["gc_geve"] = gc_of_seq(fetch_seq(
            seqs, record["contig"], record["geve_start"], record["geve_end"]
        ))
        record["tir"], record["tsd"] = _valid_original_tir(
            marker, record["original_geve_name"],
            record["geve_start"], record["geve_end"],
        )
        record["has_tir"] = record["tir"] is not None

    # Re-predict only contigs containing changed GEVEs, but always on the
    # complete contig. This is the same ORF-context methodology as findGEVE.
    changed_contigs = sorted(
        {r["contig"] for r in temp if r["action"] == "change"},
        key=_natural_key,
    )
    contig_orf_catalogue: Dict[str, List[ReviewOrf]] = {}
    for contig in changed_contigs:
        contig_orf_catalogue[contig] = predict_contig_orfs(
            contig, seqs[contig], host
        )

    all_changed_contig_orfs = [
        orf
        for contig in changed_contigs
        for orf in contig_orf_catalogue.get(contig, [])
    ]
    scan_changed_orfs(all_changed_contig_orfs, db, evalue, threads)

    for record in temp:
        old_orfs = original_orfs(
            record, marker, seqs, original_peptides=original_peptides
        )
        if record["action"] == "change":
            record["orfs"] = select_orfs_in_record(
                record, contig_orf_catalogue.get(record["contig"], [])
            )
            delta = len(record["orfs"]) - len(old_orfs)
            _LOG.info(
                f"ORF comparison | {record['original_geve_name']}: "
                f"original={len(old_orfs):,}, reviewed={len(record['orfs']):,}, "
                f"delta={delta:+,}; method=whole_contig_pyrodigal_gv"
            )
            if (
                record["geve_start"] <= (_safe_int(record["original_summary"].get("start"), record["geve_start"]) or record["geve_start"])
                and record["geve_end"] >= (_safe_int(record["original_summary"].get("end"), record["geve_end"]) or record["geve_end"])
                and delta < 0
            ):
                _LOG.warning(
                    f"{record['original_geve_name']}: extended boundaries still yielded "
                    f"fewer ORFs. This indicates a changed Pyrodigal-GV version, genome, "
                    f"or host-GFF input relative to the original findGEVE run."
                )
        else:
            record["orfs"] = old_orfs

    for record in temp:
        record["orfs"].sort(key=lambda o: (o.start, o.end, o.strand))
        hallmarks = sorted(
            {o.hallmark for o in record["orfs"] if o.hallmark},
            key=_hallmark_key,
        )
        record["hallmarks_present"] = hallmarks
        record["n_hallmarks"] = sum(1 for o in record["orfs"] if o.hallmark)
    return temp

def coding_density_v2(orfs, start, end):
    spans=[]
    for o in orfs:
        s,e=max(start,o.start),min(end,o.end)
        if s<=e: spans.append((s,e))
    merged=_merge_host_intervals(spans)
    return round(sum(e-s+1 for s,e in merged)/(end-start+1),4) if end>=start else 0.0

def load_gvog_names(db: Path):
    p=db/"gvog.complete.annot.tsv"
    if not p.is_file(): return {}
    try: df=pd.read_csv(p,sep="\t",dtype=str,keep_default_na=False)
    except Exception: return {}
    if not {"GVOG","NCVOG_descs"}.issubset(df.columns): return {}
    return {r["GVOG"]:r["NCVOG_descs"].split(" | ")[0].strip() for _,r in df.iterrows() if r["GVOG"]}

def write_v2_summary(records, path):
    rows=[]
    for g in records:
        row=dict(contig_id=g["contig"],geve_name=g["geve_id"],start=g["geve_start"],end=g["geve_end"],geve_length=g["geve_length"],gc=round(g["gc_geve"],2),total_cds=len(g["orfs"]),NCLDV_hits=sum(1 for o in g["orfs"] if o.hallmark or o.gvog),coding_density=coding_density_v2(g["orfs"],g["geve_start"],g["geve_end"]),n_hallmarks=g["n_hallmarks"],hallmarks=",".join(g["hallmarks_present"]),has_tir="yes" if g["tir"] else "no")
        row.update(tir_fields(g["tir"])); row.update(tsd_fields(g["tsd"])); rows.append(row)
    pd.DataFrame(rows,columns=["contig_id","geve_name","start","end","geve_length","gc","total_cds","NCLDV_hits","coding_density","n_hallmarks","hallmarks","has_tir","tir_length","tir_score","tir_identity_pct","tir_gaps","tsd_len","tsd_left","tsd_right","tsd_mismatch","tsd_conservation"]).to_csv(path,sep="\t",index=False)

def _fmt_ev(x): return "NA" if x is None or x==float("inf") or not math.isfinite(x) else f"{x:.2e}"
def _fmt_sc(x): return "NA" if not x or x<=0 else f"{x:.1f}"

def write_v2_outputs(records, seqs, outdir, file_prefix, db):
    summary=outdir/f"{file_prefix}.summary.tsv"; marker=outdir/f"{file_prefix}.markerout"; bed=outdir/f"{file_prefix}.geve.bed"; fna=outdir/f"{file_prefix}.geve.fna"; gff=outdir/f"{file_prefix}.geve.gff3"; cds=outdir/f"{file_prefix}.geve.cds"; pep=outdir/f"{file_prefix}.geve.pep"; func=outdir/f"{file_prefix}.func.tsv"
    write_v2_summary(records,summary)
    with fna.open("w") as fh:
        for g in records:
            fh.write(f">{g['geve_id']} contig={g['contig']} start={g['geve_start']} end={g['geve_end']} length={g['geve_length']} hallmarks={','.join(g['hallmarks_present'])} gc={g['gc_geve']:.2f}%\n{wrap_fasta(fetch_seq(seqs,g['contig'],g['geve_start'],g['geve_end']))}\n")
    with cds.open("w") as cf, pep.open("w") as pf, gff.open("w") as gf:
        gf.write("##gff-version 3\n")
        for g in records:
            gid=g["geve_id"]; gf.write(f"{g['contig']}\tfindGEVE\tmobile_genetic_element\t{g['geve_start']}\t{g['geve_end']}\t.\t+\t.\tID={gid};Name={gid}\n")
            tir = g["tir"]
            tsd = g["tsd"]
            if tir is not None:
                gf.write(f"{g['contig']}\tfindGEVE\tterminal_inverted_repeat\t{tir.left_start}\t{tir.left_end}\t.\t+\t.\tID={gid}.TIR_left;Parent={gid}\n")
                gf.write(f"{g['contig']}\tfindGEVE\tterminal_inverted_repeat\t{tir.right_start}\t{tir.right_end}\t.\t-\t.\tID={gid}.TIR_right;Parent={gid}\n")
                if tsd is not None:
                    le = tir.left_start - 1 - tsd.left_shift
                    ls = le - tsd.length + 1
                    rs = tir.right_end + 1 + tsd.right_shift
                    re = rs + tsd.length - 1
                    gf.write(f"{g['contig']}\tfindGEVE\ttarget_site_duplication\t{ls}\t{le}\t.\t+\t.\tID={gid}.TSD_5p;Parent={gid}\n")
                    gf.write(f"{g['contig']}\tfindGEVE\ttarget_site_duplication\t{rs}\t{re}\t.\t+\t.\tID={gid}.TSD_3p;Parent={gid}\n")
            for i,o in enumerate(g["orfs"],1):
                lab=f"orf{i:05d}"; strand="+" if o.strand>=0 else "-"; nt=fetch_seq(seqs,o.contig,o.start,o.end); nt=revcomp(nt) if o.strand<0 else nt; extra=f" {o.hallmark}" if o.hallmark else ""
                cf.write(f">{gid}_{lab}{extra} length={len(nt)}\n{wrap_fasta(nt)}\n"); pf.write(f">{gid}_{lab}{extra} length={len(o.protein)}\n{wrap_fasta(o.protein)}\n")
                score=f"{o.hallmark_bitscore:.1f}" if o.hallmark else "."; attrs=f"ID={gid}.{lab};Parent={gid};Name={lab}"+(f";hallmark={o.hallmark}" if o.hallmark else "")
                gf.write(f"{o.contig}\tfindGEVE\tCDS\t{o.start}\t{o.end}\t{score}\t{strand}\t0\t{attrs}\n")
    with marker.open("w") as fh:
        fh.write("contig\tgeve_name\tfeature\tname\tstart\tend\tstrand\te_value\tscore\n")
        for g in records:
            gid=g["geve_id"]; flank=viz_flank_size(g["geve_length"]); rs=max(1,g["geve_start"]-flank); re=min(len(seqs[g["contig"]]),g["geve_end"]+flank)
            fh.write(f"{g['contig']}\t{gid}\tGEVE\t.\t{g['geve_start']}\t{g['geve_end']}\t.\tNA\t{g['geve_length']}\n")
            if rs<g["geve_start"]: fh.write(f"{g['contig']}\t{gid}\tflank_left\t.\t{rs}\t{g['geve_start']-1}\t.\tNA\tNA\n")
            if re>g["geve_end"]: fh.write(f"{g['contig']}\t{gid}\tflank_right\t.\t{g['geve_end']+1}\t{re}\t.\tNA\tNA\n")
            tir = g["tir"]
            tsd = g["tsd"]
            if tir is not None:
                fh.write(f"{g['contig']}\t{gid}\tTIR_left\t.\t{tir.left_start}\t{tir.left_end}\t+\tNA\t{tir.score}\n")
                fh.write(f"{g['contig']}\t{gid}\tTIR_right\t.\t{tir.right_start}\t{tir.right_end}\t-\tNA\t{tir.score}\n")
                if tsd is not None:
                    le = tir.left_start - 1 - tsd.left_shift
                    ls = le - tsd.length + 1
                    rs = tir.right_end + 1 + tsd.right_shift
                    re2 = rs + tsd.length - 1
                    fh.write(f"{g['contig']}\t{gid}\tTSD_5p\t{tsd.sequence_left}\t{ls}\t{le}\t+\tNA\t{tsd.identity:.1f}\n")
                    fh.write(f"{g['contig']}\t{gid}\tTSD_3p\t{tsd.sequence_right}\t{rs}\t{re2}\t+\tNA\t{tsd.identity:.1f}\n")
            for o in g["orfs"]:
                strand="+" if o.strand>=0 else "-"
                if o.hallmark: feat,name,ev,sc="hallmark",o.hallmark,_fmt_ev(o.hallmark_evalue),_fmt_sc(o.hallmark_bitscore)
                elif o.gvog: feat,name,ev,sc="gvog",o.gvog,_fmt_ev(o.gvog_evalue),_fmt_sc(o.gvog_bitscore)
                elif o.best_pfam_acc: feat,name,ev,sc="pfam",o.best_pfam_name or o.best_pfam_acc,_fmt_ev(o.best_pfam_evalue),_fmt_sc(o.best_pfam_bitscore)
                else: feat,name,ev,sc="orf",".","NA","NA"
                fh.write(f"{o.contig}\t{gid}\t{feat}\t{name}\t{o.start}\t{o.end}\t{strand}\t{ev}\t{sc}\n")
    rows=[]
    names=load_gvog_names(db)
    for g in records:
        for i,o in enumerate(g["orfs"],1):
            if not (o.hallmark or o.gvog or o.best_pfam_acc): continue
            rows.append(dict(geve_id=g["geve_id"],protein_id=f"orf{i:05d}",gvog_bitscore=_fmt_sc(o.gvog_bitscore),gvog_evalue=_fmt_ev(o.gvog_evalue),gvog_id=o.gvog or "NA",gvog_name=names.get(o.gvog,"NA") if o.gvog else "NA",pfam_bitscore=_fmt_sc(o.best_pfam_bitscore),pfam_evalue=_fmt_ev(o.best_pfam_evalue),pfam_id=o.best_pfam_acc or "NA",pfam_name=o.best_pfam_name or "NA"))
    pd.DataFrame(rows,columns=["geve_id","protein_id","gvog_bitscore","gvog_evalue","gvog_id","gvog_name","pfam_bitscore","pfam_evalue","pfam_id","pfam_name"]).to_csv(func,sep="\t",index=False)
    byh=defaultdict(dict)
    for g in records:
        for o in g["orfs"]:
            if o.hallmark and (g["geve_id"] not in byh[o.hallmark] or len(o.protein)>len(byh[o.hallmark][g["geve_id"]].protein)): byh[o.hallmark][g["geve_id"]]=o
    hdir=outdir/"hallmark"
    if hdir.is_dir():
        for old in hdir.glob(f"{file_prefix}.*.pep"):
            old.unlink()
    if byh:
        hdir.mkdir(exist_ok=True)
        for h in sorted(byh,key=_hallmark_key):
            with (hdir/f"{file_prefix}.{h}.pep").open("w") as fh:
                for gid,o in sorted(byh[h].items(),key=lambda x:_natural_key(x[0])): fh.write(f">{gid}_{h}\n{wrap_fasta(o.protein)}\n")
    elif hdir.is_dir() and not any(hdir.iterdir()):
        hdir.rmdir()
    rolling={}
    for g in records:
        vals={}
        a=g["orfs"]
        for i,o in enumerate(a):
            lo=max(0,i-7); hi=min(len(a),i+8); vals[o.orf_id]=sum(x.net_score for x in a[lo:hi])/(hi-lo) if hi-lo>=3 else 0.0
        rolling[g["geve_id"]]=vals
    rows=[]
    for g in records:
        flank=viz_flank_size(g["geve_length"]); rs=max(1,g["geve_start"]-flank); re=min(len(seqs[g["contig"]]),g["geve_end"]+flank); w=rs
        while w+999<=re:
            we=w+999; sub=fetch_seq(seqs,g["contig"],w,we); gc=gc_of_seq(sub); ovs=[o for o in g["orfs"] if o.end>=w and o.start<=we]; rv=[rolling[g["geve_id"]][o.orf_id] for o in ovs]
            center=(w+we)//2; rt="flank_left" if center<g["geve_start"] else ("geve" if center<=g["geve_end"] else "flank_right")
            rows.append(dict(contig_id=g["contig"], window_start=w, window_end=we, geve_name=g["geve_id"], rel_start=w-g["geve_start"], rel_end=we-g["geve_start"], region_type=rt, gc="NA" if math.isnan(gc) else f"{gc:.3f}", rolling_score_mean="NA" if not rv else f"{sum(rv)/len(rv):.4f}", n_orfs=len(ovs), gvog_hits=sum(1 for o in ovs if o.hallmark or o.gvog), pfam_hits=sum(1 for o in ovs if o.best_pfam_acc)))
            w += 250
    pd.DataFrame(rows,columns=["contig_id","window_start","window_end","geve_name","rel_start","rel_end","region_type","gc","rolling_score_mean","n_orfs","gvog_hits","pfam_hits"]).to_csv(bed,sep="\t",index=False)
    return summary,marker,bed

def validate_database(db: Path) -> None:
    if not db.is_dir():
        raise SystemExit(f"Error: database directory not found: {db}")
    for name in ("NCLDV_markers.hmm", "gvog.complete.hmm"):
        path = db / name
        if not path.is_file():
            raise SystemExit(f"Error: required database file not found: {path}")


def default_region_outdir(base: Optional[Path]) -> Path:
    root = base if base is not None else Path.cwd()
    date_tag = datetime.now().strftime("%Y%m%d")
    out = root / f"Result_{date_tag}"
    if not out.exists():
        return out
    idx = 1
    while True:
        candidate = root / f"Result_{date_tag}_{idx:02d}"
        if not candidate.exists():
            return candidate
        idx += 1


def prepare_outdir(outdir: Path, overwrite: bool) -> None:
    if outdir.exists() and any(outdir.iterdir()) and not overwrite:
        raise SystemExit(f"Error: output directory is not empty; use --overwrite: {outdir}")
    outdir.mkdir(parents=True, exist_ok=True)


def apply_review(args) -> None:
    summary = _read_table(args.summary, "summary")
    marker = _read_table(args.markerout, "markerout")
    review,errors,warnings=validate_review(read_review_xlsx(args.review),summary)
    prefix=args.prefix or infer_prefix(args.summary,summary); outdir=args.outdir or default_outdir(args.outbase)
    prepare_outdir(outdir, args.overwrite); setup_logging(outdir/"review.log")
    if errors:
        for x in errors:_LOG.error(x)
        raise SystemExit(f"Error: review validation failed with {len(errors)} error(s)")
    for x in warnings:_LOG.warning(x)
    if args.genome is None: raise SystemExit("Error: --genome is required for review reannotation")
    validate_database(args.db)
    _require_columns(marker, ["contig", "geve_name", "feature", "name", "start", "end", "strand", "e_value", "score"], "markerout")
    seqs = read_fasta(args.genome)
    host = parse_host_gff(args.gff)
    original_pep_path = args.original_pep
    if original_pep_path is not None and not original_pep_path.is_file():
        raise SystemExit(f"Error: original peptide FASTA not found: {original_pep_path}")
    if original_pep_path is None:
        inferred = args.summary.parent / f"{infer_prefix(args.summary, summary)}.geve.pep"
        if inferred.is_file():
            original_pep_path = inferred
    original_peptides = read_peptide_fasta(original_pep_path)
    if original_peptides:
        _LOG.info(
            f"Loaded {len(original_peptides):,} original peptide(s) from "
            f"{original_pep_path}; unchanged GEVEs will preserve exact peptide sequences"
        )
    else:
        _LOG.warning(
            "Original GEVE peptide FASTA was not found; unchanged peptides will be "
            "reconstructed from genomic coordinates and may differ at alternative starts"
        )
    records = build_v2_records(
        review, summary, marker, seqs, host, args.db, args.evalue,
        args.threads, prefix, original_peptides=original_peptides,
    )
    file_prefix=f"{prefix}.reviewed"; sp,mp,bp=write_v2_outputs(records,seqs,outdir,file_prefix,args.db)
    if not args.no_plot: run_geve_plot(mp,bp,outdir)
    _LOG.info(f"Retained {len(records):,} GEVE(s); removed {int((review['action']=='remove').sum()):,}; reannotated {int((review['action']=='change').sum()):,}")
    if (outdir / "hallmark").is_dir():
        _LOG.output(f"Wrote hallmark protein folder: {outdir / 'hallmark'}")
    else:
        _LOG.info("No hallmark proteins detected; hallmark output directory was not created")
    _LOG.output(f"Wrote {sp}")



def _parse_region_values(values: List[str], label: str, integer: bool = False):
    text = " ".join(str(value) for value in values).strip()
    parts = [part.strip() for part in text.split(",")]
    if not parts or any(not part for part in parts):
        raise SystemExit(f"Error: {label} contains an empty value")
    if not integer:
        return parts
    parsed = []
    for part in parts:
        try:
            parsed.append(int(part))
        except ValueError as exc:
            raise SystemExit(f"Error: {label} contains a non-integer value: {part}") from exc
    return parsed


def inspect_region(args) -> None:
    prefix = str(args.prefix).strip()
    if not prefix:
        raise SystemExit("Error: --prefix must not be empty")
    if args.threads < 1:
        raise SystemExit("Error: --threads must be at least 1")
    if args.evalue <= 0:
        raise SystemExit("Error: --evalue must be greater than 0")

    contigs = _parse_region_values(args.ctg, "--ctg")
    starts = _parse_region_values(args.start, "--start", integer=True)
    ends = _parse_region_values(args.end, "--end", integer=True)
    if not (len(contigs) == len(starts) == len(ends)):
        raise SystemExit(
            "Error: --ctg, --start, and --end must contain the same number of values "
            f"(received {len(contigs)}, {len(starts)}, and {len(ends)})"
        )

    validate_database(args.db)
    outdir = args.outdir or default_region_outdir(args.outbase)
    prepare_outdir(outdir, args.overwrite)
    setup_logging(outdir / "region.log")
    _LOG.info("findGEVE region inspection started")
    _LOG.info(f"Genome file | {args.genome}")
    _LOG.info(f"Database directory | {args.db}")
    _LOG.info(f"Candidate regions | {len(contigs):,}")
    _LOG.info(f"Output directory | {outdir}")

    seqs = read_fasta(args.genome)
    regions = []
    for index, (contig, start, end) in enumerate(zip(contigs, starts, ends), 1):
        if contig not in seqs:
            raise SystemExit(f"Error: candidate {index}: contig not found in genome: {contig}")
        contig_length = len(seqs[contig])
        if start < 1 or end > contig_length or start >= end:
            raise SystemExit(
                f"Error: candidate {index}: invalid region coordinates: "
                f"{contig}:{start}-{end} (contig length {contig_length})"
            )
        regions.append((contig, start, end))
        _LOG.info(f"Candidate {index:03d} | {contig}:{start}-{end}")

    host = parse_host_gff(args.gff)
    contig_catalogue = {
        contig: predict_contig_orfs(contig, seqs[contig], host)
        for contig in sorted(set(contigs), key=_natural_key)
    }
    all_orfs = [
        orf for contig in sorted(contig_catalogue, key=_natural_key)
        for orf in contig_catalogue[contig]
    ]
    scan_changed_orfs(all_orfs, args.db, args.evalue, args.threads)

    records = []
    for index, (contig, start, end) in enumerate(regions, 1):
        geve_id = f"{prefix}_GEVE_{index:03d}"
        record = dict(
            original_geve_name=geve_id,
            reviewed_geve_name=geve_id,
            geve_id=geve_id,
            action="region",
            contig=contig,
            geve_start=start,
            geve_end=end,
            geve_length=end - start + 1,
            tir=None,
            tsd=None,
            has_tir=False,
        )
        record["gc_geve"] = gc_of_seq(fetch_seq(seqs, contig, start, end))
        record["orfs"] = select_orfs_in_record(record, contig_catalogue[contig])
        records.append(record)
    for record in records:
        record["orfs"].sort(key=lambda o: (o.start, o.end, o.strand))
        record["hallmarks_present"] = sorted(
            {o.hallmark for o in record["orfs"] if o.hallmark},
            key=_hallmark_key,
        )
        record["n_hallmarks"] = sum(1 for o in record["orfs"] if o.hallmark)

    summary_path, marker_path, bed_path = write_v2_outputs(records, seqs, outdir, prefix, args.db)
    if not args.no_plot:
        run_geve_plot(marker_path, bed_path, outdir)
    for record in records:
        _LOG.info(
            f"Candidate retained: {record['geve_id']} | {record['contig']}:"
            f"{record['geve_start']}-{record['geve_end']} | "
            f"ORFs={len(record['orfs']):,} | hallmarks={record['n_hallmarks']:,}"
        )
    if (outdir / "hallmark").is_dir():
        _LOG.output(f"Wrote hallmark protein folder: {outdir / 'hallmark'}")
    else:
        _LOG.info("No hallmark proteins detected; hallmark output directory was not created")
    _LOG.output(f"Wrote {summary_path}")


def _program_name() -> str:
    return Path(sys.argv[0]).name or "findGEVE_review.py"


def _render_help(text: str) -> str:
    return text.format(prog=_program_name())


class _Parser(argparse.ArgumentParser):
    def __init__(self, help_text: str, usage_text: str, **kwargs):
        self._help_text = help_text
        self._usage_text = usage_text
        super().__init__(add_help=False, **kwargs)
        self.add_argument("-h", "--help", action="help", help=argparse.SUPPRESS)

    def format_help(self) -> str:
        return _render_help(self._help_text)

    def format_usage(self) -> str:
        return _render_help(self._usage_text)


def _command_parser(command: str) -> _Parser:
    if command == "make-template":
        parser = _Parser(
            MAKE_TEMPLATE_HELP_TEXT,
            "Usage: {prog} make-template <prefix.summary.tsv> [OPTIONS]\n",
            prog=f"{_program_name()} make-template",
        )
        parser.add_argument("summary", type=Path)
        parser.add_argument("--overwrite", action="store_true")
        return parser

    if command == "apply":
        parser = _Parser(
            APPLY_HELP_TEXT,
            f"Usage: {_program_name()} apply -db <directory> --genome genome.fa "
            "--review <prefix.review.xlsx> --summary <prefix.summary.tsv> "
            "--markerout <prefix.markerout> [OPTIONS]\n",
            prog=f"{_program_name()} apply",
        )
        parser.add_argument("--review", required=True, type=Path)
        parser.add_argument("--summary", required=True, type=Path)
        parser.add_argument("--markerout", required=True, type=Path)
        parser.add_argument("--genome", required=True, type=Path)
        parser.add_argument("-db", "--db", required=True, type=Path)
        parser.add_argument("-g", "--gff", type=Path)
        parser.add_argument("--original-pep", type=Path)
        parser.add_argument("-e", "--evalue", type=float, default=1e-5)
        parser.add_argument("-t", "--threads", type=int, default=4)
        parser.add_argument("--prefix")
        parser.add_argument("-o", "--outdir", type=Path)
        parser.add_argument("--outbase", type=Path)
        parser.add_argument("--overwrite", action="store_true")
        parser.add_argument("--no-plot", action="store_true")
        return parser

    if command == "region":
        parser = _Parser(
            REGION_HELP_TEXT,
            f"Usage: {_program_name()} region -db <directory> --genome genome.fa "
            "--prefix <prefix> --ctg <contig[,contig...]> "
            "--start <start[,start...]> --end <end[,end...]> [OPTIONS]\n",
            prog=f"{_program_name()} region",
        )
        parser.add_argument("--genome", required=True, type=Path)
        parser.add_argument("-db", "--db", required=True, type=Path)
        parser.add_argument("--prefix", required=True)
        parser.add_argument("--ctg", required=True, nargs="+")
        parser.add_argument("--start", required=True, nargs="+")
        parser.add_argument("--end", required=True, nargs="+")
        parser.add_argument("-g", "--gff", type=Path)
        parser.add_argument("-e", "--evalue", type=float, default=1e-5)
        parser.add_argument("-t", "--threads", type=int, default=4)
        parser.add_argument("-o", "--outdir", type=Path)
        parser.add_argument("--outbase", type=Path)
        parser.add_argument("--overwrite", action="store_true")
        parser.add_argument("--no-plot", action="store_true")
        return parser

    raise ValueError(command)


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    values = list(sys.argv[1:] if argv is None else argv)
    if not values or values[0] in {"-h", "--help"}:
        sys.stdout.write(_render_help(HELP_TEXT))
        raise SystemExit(0)

    command = values.pop(0)
    if command not in {"make-template", "apply", "region"}:
        sys.stderr.write(f"Error: unknown command: {command}\n\n")
        sys.stderr.write(_render_help(HELP_TEXT))
        raise SystemExit(2)

    args = _command_parser(command).parse_args(values)
    args.command = command
    return args


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    if args.command == "make-template":
        setup_logging()
        make_template(args.summary, overwrite=args.overwrite)
        return 0
    if args.command == "apply":
        setup_logging()
        if args.threads < 1:
            raise SystemExit("Error: --threads must be at least 1")
        if args.evalue <= 0:
            raise SystemExit("Error: --evalue must be greater than 0")
        apply_review(args)
        return 0
    if args.command == "region":
        inspect_region(args)
        return 0
    return 2

if __name__ == "__main__":
    sys.exit(main())